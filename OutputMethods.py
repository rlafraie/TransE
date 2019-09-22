from openpyxl import Workbook, load_workbook
from pathlib import Path
from typing import List
from matplotlib import pyplot as plt


def create_hyper_param_sheet(kg_log_dir: Path, title: str, sheet_file_name: str, header: List[str]):
    wb = Workbook()
    ws = wb.active
    ws.title = title

    ws.append(header)
    # ws.add_table(tab)
    wb.save(kg_log_dir / sheet_file_name)
    wb.close()


def update_hyper_param_sheet(kg_log_dir: Path, sheet_file_name: str, input_row: List):
    wb = load_workbook(kg_log_dir / sheet_file_name)
    ws = wb.active

    ws.append(input_row)
    wb.save(kg_log_dir / sheet_file_name)
    wb.close()


def initialize_log_folder(knowledge_graph_dir: Path):
    kg_log_dir = knowledge_graph_dir / 'evaluation_earlyStop'
    if kg_log_dir.exists():
        max_hyper_param_id = len([folder.name for folder in kg_log_dir.iterdir() if folder.is_dir()])
        hyper_param_config_dir = kg_log_dir / str(max_hyper_param_id + 1)
        hyper_param_config_dir.mkdir()
    else:
        hyper_param_config_dir = kg_log_dir / '1'
        hyper_param_config_dir.mkdir(parents=True)
        create_hyper_param_sheet(kg_log_dir, 'Hyperparameter configurations', 'hyper_param_mapping.xlsx'
                                 , ['hyper_param_id', 'num_of_epochs', 'batch_size', 'margin', 'norm',
                                    'learning_rate', 'num_of_dimensions', 'learned_epochs'])

        create_hyper_param_sheet(kg_log_dir, 'Evaluation Scores', 'hyper_param_scores.xlsx'
                                 , ['hyper_param_id', 'raw Validation MR', 'filtered Validation MR',
                                    'raw Validation Hits@10',
                                    'filtered Validation Hits@10', 'raw Test MR', 'filtered Test MR',
                                    'raw Test Hits@10',
                                    'filtered Test Hits@10'])

    return hyper_param_config_dir


def save_figure(hyper_param_path: Path, filename: str, title: str, xlabel: str, ylabel: str,
                training_data_points: List, validation_data_points: List, losses: List, num_of_epochs, validation_freq: int):
    lines = []

    plt.grid(True)
    plt.title(title)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)

    training_line, = plt.plot(range(1, len(training_data_points) + 1), training_data_points, 'g',
                              label='Training_score')
    lines.append(training_line)

    if validation_data_points:
        x_points = [i for i in range(validation_freq, num_of_epochs, validation_freq)] + [num_of_epochs]
        plt.xticks(range(1, len(x_points) + 1), x_points)

        valid_line, = plt.plot(range(1, len(validation_data_points) + 1), validation_data_points, 'b',
                               label='Validation_Score')
        lines.append(valid_line)

    else:
        x_points = [1] + [i for i in range(validation_freq, num_of_epochs, validation_freq)] + [num_of_epochs]
        plt.xticks(x_points)

    if losses:
        losses_line, = plt.plot(range(1, len(losses) + 1), losses, 'k', label='Loss')
        lines.append(losses_line)

    plt.legend(handles=lines)
    plt.savefig(hyper_param_path / filename, quality=95)
    plt.close()
